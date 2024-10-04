# Copyright 2021-2024 Avaiga Private Limited
#
# Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
# the License. You may obtain a copy of the License at
#
#        http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
# an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
# specific language governing permissions and limitations under the License.

from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Set, Union

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from taipy.common.config.common.scope import Scope

from .._entity._reload import _Reloader
from .._version._version_manager_factory import _VersionManagerFactory
from ..exceptions.exceptions import ExposedTypeLengthMismatch, NonExistingExcelSheet, SheetNameLengthMismatch
from ..job.job_id import JobId
from ._file_datanode_mixin import _FileDataNodeMixin
from ._tabular_datanode_mixin import _TabularDataNodeMixin
from .data_node import DataNode
from .data_node_id import DataNodeId, Edit


class ExcelDataNode(DataNode, _FileDataNodeMixin, _TabularDataNodeMixin):
    """Data Node stored as an Excel file.

    The Excel file format is _xlsx_.

    The *properties* attribute can contain the following optional entries:

    - *sheet_name* (`Union[str, List[str]]`): The name of the sheet(s) to be used.
    - *default_path* (`str`): The default path of the Excel file used at the instantiation of the
        data node.
    - *default_data*: The default data of the data node. It is used at the data node instantiation
        to write the data to the Excel file.
    - *has_header* (`bool`): If True, indicates that the Excel file has a header.
    - *exposed_type* (`str`): The exposed type of the data read from Excel file. The default value
        is `pandas`.
    """

    __STORAGE_TYPE = "excel"
    __SHEET_NAME_PROPERTY = "sheet_name"

    _REQUIRED_PROPERTIES: List[str] = []

    def __init__(
        self,
        config_id: str,
        scope: Scope,
        id: Optional[DataNodeId] = None,
        owner_id: Optional[str] = None,
        parent_ids: Optional[Set[str]] = None,
        last_edit_date: Optional[datetime] = None,
        edits: List[Edit] = None,
        version: str = None,
        validity_period: Optional[timedelta] = None,
        edit_in_progress: bool = False,
        editor_id: Optional[str] = None,
        editor_expiration_date: Optional[datetime] = None,
        properties: Dict = None,
    ) -> None:
        self.id = id or self._new_id(config_id)

        if properties is None:
            properties = {}

        if self.__SHEET_NAME_PROPERTY not in properties.keys():
            properties[self.__SHEET_NAME_PROPERTY] = None
        if self._HAS_HEADER_PROPERTY not in properties.keys():
            properties[self._HAS_HEADER_PROPERTY] = True
        properties[self._EXPOSED_TYPE_PROPERTY] = _TabularDataNodeMixin._get_valid_exposed_type(properties)
        self._check_exposed_type(properties[self._EXPOSED_TYPE_PROPERTY])

        default_value = properties.pop(self._DEFAULT_DATA_KEY, None)
        _FileDataNodeMixin.__init__(self, properties)
        _TabularDataNodeMixin.__init__(self, **properties)

        DataNode.__init__(
            self,
            config_id,
            scope,
            self.id,
            owner_id,
            parent_ids,
            last_edit_date,
            edits,
            version or _VersionManagerFactory._build_manager()._get_latest_version(),
            validity_period,
            edit_in_progress,
            editor_id,
            editor_expiration_date,
            **properties,
        )

        with _Reloader():
            self._write_default_data(default_value)

        self._TAIPY_PROPERTIES.update(
            {
                self._PATH_KEY,
                self._DEFAULT_PATH_KEY,
                self._DEFAULT_DATA_KEY,
                self._IS_GENERATED_KEY,
                self._HAS_HEADER_PROPERTY,
                self._EXPOSED_TYPE_PROPERTY,
                self.__SHEET_NAME_PROPERTY,
            }
        )

    @classmethod
    def storage_type(cls) -> str:
        """Return the storage type of the data node: "excel"."""
        return cls.__STORAGE_TYPE

    def write_with_column_names(self, data: Any, columns: List[str] = None, job_id: Optional[JobId] = None) -> None:
        """Write a set of columns.

        Parameters:
            data (Any): The data to write.
            columns (List[str]): The list of column names to write.
            job_id (Optional[JobId]): An optional identifier of the writer.
        """
        if isinstance(data, Dict) and all(isinstance(x, (pd.DataFrame, np.ndarray)) for x in data.values()):
            self._write_excel_with_multiple_sheets(data, columns=columns)
        else:
            df = pd.DataFrame(data)
            if columns:
                df = self._set_column_if_dataframe(df, columns)
            self._write_excel_with_single_sheet(df.to_excel, self.path, index=False)
        self.track_edit(timestamp=datetime.now(), job_id=job_id)

    @staticmethod
    def _check_exposed_type(exposed_type):
        if isinstance(exposed_type, str):
            _TabularDataNodeMixin._check_exposed_type(exposed_type)
        elif isinstance(exposed_type, list):
            for t in exposed_type:
                _TabularDataNodeMixin._check_exposed_type(t)
        elif isinstance(exposed_type, dict):
            for t in exposed_type.values():
                _TabularDataNodeMixin._check_exposed_type(t)

    def _read(self):
        return self._read_from_path()

    def _read_from_path(self, path: Optional[str] = None, **read_kwargs) -> Any:
        if path is None:
            path = self._path

        exposed_type = self.properties[self._EXPOSED_TYPE_PROPERTY]
        if exposed_type == self._EXPOSED_TYPE_PANDAS:
            return self._read_as_pandas_dataframe(path=path)
        if exposed_type == self._EXPOSED_TYPE_NUMPY:
            return self._read_as_numpy(path=path)
        return self._read_as(path=path)

    def _read_sheet_with_exposed_type(
        self, path: str, sheet_exposed_type: str, sheet_name: str
    ) -> Optional[Union[np.ndarray, pd.DataFrame]]:
        if sheet_exposed_type == self._EXPOSED_TYPE_NUMPY:
            return self._read_as_numpy(path, sheet_name)
        elif sheet_exposed_type == self._EXPOSED_TYPE_PANDAS:
            return self._read_as_pandas_dataframe(path, sheet_name)  # type: ignore
        return None

    def _read_as(self, path: str):
        try:
            properties = self.properties
            excel_file = load_workbook(path, read_only=True)
            exposed_type = properties[self._EXPOSED_TYPE_PROPERTY]
            work_books = {}
            sheet_names = excel_file.sheetnames

            user_provided_sheet_names = properties.get(self.__SHEET_NAME_PROPERTY) or []
            if not isinstance(user_provided_sheet_names, (list, set, tuple)):
                user_provided_sheet_names = [user_provided_sheet_names]

            provided_sheet_names = user_provided_sheet_names or sheet_names

            for sheet_name in provided_sheet_names:
                if sheet_name not in sheet_names:
                    raise NonExistingExcelSheet(sheet_name, path)

            if isinstance(exposed_type, List):
                if len(provided_sheet_names) != len(exposed_type):
                    raise ExposedTypeLengthMismatch(
                        f"Expected {len(provided_sheet_names)} exposed types, got " f"{len(exposed_type)}"
                    )

            for i, sheet_name in enumerate(provided_sheet_names):
                work_sheet = excel_file[sheet_name]
                sheet_exposed_type = exposed_type

                if not isinstance(sheet_exposed_type, str):
                    if isinstance(exposed_type, dict):
                        sheet_exposed_type = exposed_type.get(sheet_name, self._EXPOSED_TYPE_PANDAS)
                    elif isinstance(exposed_type, List):
                        sheet_exposed_type = exposed_type[i]

                    if isinstance(sheet_exposed_type, str):
                        sheet_data = self._read_sheet_with_exposed_type(path, sheet_exposed_type, sheet_name)
                        if sheet_data is not None:
                            work_books[sheet_name] = sheet_data
                        continue

                res = [[col.value for col in row] for row in work_sheet.rows]
                if properties[self._HAS_HEADER_PROPERTY] and res:
                    header = res.pop(0)
                    for i, row in enumerate(res):
                        res[i] = sheet_exposed_type(**dict([[h, r] for h, r in zip(header, row)]))
                else:
                    for i, row in enumerate(res):
                        res[i] = sheet_exposed_type(*row)
                work_books[sheet_name] = res  # type: ignore
        finally:
            excel_file.close()

        if len(user_provided_sheet_names) == 1:
            return work_books[user_provided_sheet_names[0]]

        return work_books

    def _read_as_numpy(self, path: str, sheet_names=None):
        sheets = self._read_as_pandas_dataframe(path=path, sheet_names=sheet_names)
        if isinstance(sheets, dict):
            return {sheet_name: df.to_numpy() for sheet_name, df in sheets.items()}
        return sheets.to_numpy()

    def _do_read_excel(
        self, path: str, sheet_names, kwargs
    ) -> Union[Dict[Union[int, str], pd.DataFrame], pd.DataFrame]:
        return pd.read_excel(path, sheet_name=sheet_names, **kwargs)

    def __get_sheet_names_and_header(self, sheet_names):
        kwargs = {}
        properties = self.properties
        if sheet_names is None:
            sheet_names = properties[self.__SHEET_NAME_PROPERTY]
        if not properties[self._HAS_HEADER_PROPERTY]:
            kwargs["header"] = None
        return sheet_names, kwargs

    def _read_as_pandas_dataframe(
        self, path: str, sheet_names=None
    ) -> Union[Dict[Union[int, str], pd.DataFrame], pd.DataFrame]:
        sheet_names, kwargs = self.__get_sheet_names_and_header(sheet_names)
        try:
            return self._do_read_excel(path, sheet_names, kwargs)
        except pd.errors.EmptyDataError:
            return pd.DataFrame()

    def _append_excel_with_single_sheet(self, append_excel_fct, *args, **kwargs):
        sheet_name = self.properties.get(self.__SHEET_NAME_PROPERTY)

        with pd.ExcelWriter(self._path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            if sheet_name:
                if not isinstance(sheet_name, str):
                    sheet_name = sheet_name[0]
                append_excel_fct(
                    writer, *args, **kwargs, sheet_name=sheet_name, startrow=writer.sheets[sheet_name].max_row
                )
            else:
                sheet_name = list(writer.sheets.keys())[0]
                append_excel_fct(writer, *args, **kwargs, startrow=writer.sheets[sheet_name].max_row)

    def _set_column_if_dataframe(self, data: Any, columns) -> Union[pd.DataFrame, Any]:
        if isinstance(data, pd.DataFrame):
            data.columns = pd.Index(columns, dtype="object")
        return data

    def _append_excel_with_multiple_sheets(self, data: Any, columns: List[str] = None):
        with pd.ExcelWriter(self._path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            # Each key stands for a sheet name
            for sheet_name in data.keys():
                if isinstance(data[sheet_name], np.ndarray):
                    df = pd.DataFrame(data[sheet_name])
                else:
                    df = data[sheet_name]

                if columns:
                    df = self._set_column_if_dataframe(df, columns)

                df.to_excel(
                    writer, sheet_name=sheet_name, index=False, header=False, startrow=writer.sheets[sheet_name].max_row
                )

    def _append(self, data: Any):
        from importlib.metadata import version

        if version("pandas") < "1.4":
            raise ImportError("The append method is only available for pandas version 1.4 or higher.")

        if isinstance(data, Dict) and all(isinstance(x, (pd.DataFrame, np.ndarray)) for x in data.values()):
            self._append_excel_with_multiple_sheets(data)
        elif isinstance(data, pd.DataFrame):
            self._append_excel_with_single_sheet(data.to_excel, index=False, header=False)
        else:
            self._append_excel_with_single_sheet(pd.DataFrame(data).to_excel, index=False, header=False)

    def _write_excel_with_single_sheet(self, write_excel_fct, *args, **kwargs):
        if sheet_name := self.properties.get(self.__SHEET_NAME_PROPERTY):
            if not isinstance(sheet_name, str):
                if len(sheet_name) > 1:
                    raise SheetNameLengthMismatch
                else:
                    sheet_name = sheet_name[0]
            write_excel_fct(*args, **kwargs, sheet_name=sheet_name)
        else:
            write_excel_fct(*args, **kwargs)

    def _write_excel_with_multiple_sheets(self, data: Any, columns: List[str] = None):
        with pd.ExcelWriter(self._path) as writer:
            # Each key stands for a sheet name
            properties = self.properties
            for key in data.keys():
                df = self._convert_data_to_dataframe(properties[self._EXPOSED_TYPE_PROPERTY], data[key])

                if columns:
                    df = self._set_column_if_dataframe(df, columns)

                df.to_excel(writer, key, index=False, header=properties[self._HAS_HEADER_PROPERTY] or False)

    def _write(self, data: Any):
        if isinstance(data, Dict):
            return self._write_excel_with_multiple_sheets(data)
        else:
            properties = self.properties
            data = self._convert_data_to_dataframe(properties[self._EXPOSED_TYPE_PROPERTY], data)
            self._write_excel_with_single_sheet(
                data.to_excel, self._path, index=False, header=properties[self._HAS_HEADER_PROPERTY] or None
            )
