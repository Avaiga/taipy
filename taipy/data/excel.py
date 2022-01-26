import csv
from collections import defaultdict
from datetime import datetime
from os.path import isfile
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
from openpyxl import load_workbook

from taipy.common.alias import DataSourceId, JobId
from taipy.data.data_source import DataSource
from taipy.data.scope import Scope
from taipy.exceptions import MissingRequiredProperty
from taipy.exceptions.data_source import NonExistingExcelSheet


class ExcelDataSource(DataSource):
    """
    A Data Source stored as an Excel file (xlsx format).

    Attributes:
        config_name (str):  Name that identifies the data source.
            We strongly recommend to use lowercase alphanumeric characters, dash character '-', or underscore character
            '_'. Note that other characters are replaced according the following rules :
            - Space characters are replaced by underscore characters ('_').
            - Unicode characters are replaced by a corresponding alphanumeric character using the Unicode library.
            - Other characters are replaced by dash characters ('-').
        scope (Scope):  The usage scope of this data source.
        id (str): Unique identifier of this data source.
        name (str): User-readable name of the data source.
        parent_id (str): Identifier of the parent (pipeline_id, scenario_id, cycle_id) or `None`.
        last_edition_date (datetime):  Date and time of the last edition.
        job_ids (List[str]): Ordered list of jobs that have written this data source.
        up_to_date (bool): `True` if the data is considered as up to date. `False` otherwise.
        properties (list): List of additional arguments. Note that the properties parameter should at least contain
           a value for "path" properties.
    """

    __STORAGE_TYPE = "excel"
    __EXPOSED_TYPE_PROPERTY = "exposed_type"
    __REQUIRED_PATH_PROPERTY = "path"
    __HAS_HEADER_PROPERTY = "has_header"
    __SHEET_NAME_PROPERTY = "sheet_name"
    __DEFAULT_SHEET_NAME = "Sheet1"
    REQUIRED_PROPERTIES = [__REQUIRED_PATH_PROPERTY]

    def __init__(
        self,
        config_name: str,
        scope: Scope,
        id: Optional[DataSourceId] = None,
        name: Optional[str] = None,
        parent_id: Optional[str] = None,
        last_edition_date: Optional[datetime] = None,
        job_ids: List[JobId] = None,
        validity_days: Optional[int] = None,
        validity_hours: Optional[int] = None,
        validity_minutes: Optional[int] = None,
        edition_in_progress: bool = False,
        properties: Dict = None,
    ):
        if properties is None:
            properties = {}
        if missing := set(self.REQUIRED_PROPERTIES) - set(properties.keys()):
            raise MissingRequiredProperty(
                f"The following properties " f"{', '.join(x for x in missing)} were not informed and are required"
            )
        if self.__SHEET_NAME_PROPERTY not in properties.keys():
            properties[self.__SHEET_NAME_PROPERTY] = self.__DEFAULT_SHEET_NAME
        if self.__HAS_HEADER_PROPERTY not in properties.keys():
            properties[self.__HAS_HEADER_PROPERTY] = True

        super().__init__(
            config_name,
            scope,
            id,
            name,
            parent_id,
            last_edition_date,
            job_ids,
            validity_days,
            validity_hours,
            validity_minutes,
            edition_in_progress,
            **properties,
        )
        if not self.last_edition_date and isfile(self.properties[self.__REQUIRED_PATH_PROPERTY]):
            self.unlock_edition()

    @classmethod
    def storage_type(cls) -> str:
        return cls.__STORAGE_TYPE

    def _read(self):
        if self.__EXPOSED_TYPE_PROPERTY in self.properties:
            return self._read_as(self.properties[self.__EXPOSED_TYPE_PROPERTY])
        try:
            return self._read_as_pandas_dataframe()
        except pd.errors.EmptyDataError:
            return pd.DataFrame()

    def _read_as(self, custom_class):
        excel_file = load_workbook(self.properties[self.__REQUIRED_PATH_PROPERTY])

        sheet_names = self.properties[self.__SHEET_NAME_PROPERTY]
        sheet_names = sheet_names if isinstance(sheet_names, (List, Set, Tuple)) else [sheet_names]
        work_books = defaultdict()

        for sheet_name in sheet_names:
            if not (sheet_name in excel_file.sheetnames):
                raise NonExistingExcelSheet(sheet_name, self.properties[self.__REQUIRED_PATH_PROPERTY])

            work_sheet = excel_file[sheet_name]
            res = list()
            for row in work_sheet.rows:
                res.append([col.value for col in row])
            if self.properties[self.__HAS_HEADER_PROPERTY]:
                header = res.pop(0)
                for i, row in enumerate(res):
                    res[i] = custom_class(**dict([[h, r] for h, r in zip(header, row)]))
            else:
                for i, row in enumerate(res):
                    res[i] = custom_class(*row)
            work_books[sheet_name] = res

        if len(sheet_names) == 1:
            return work_books[sheet_names[0]]

        return work_books

    def _read_as_pandas_dataframe(self, usecols: Optional[List[int]] = None, column_names: Optional[List[str]] = None):
        if self.properties[self.__HAS_HEADER_PROPERTY]:
            if column_names:
                return pd.read_excel(
                    self.properties[self.__REQUIRED_PATH_PROPERTY],
                    sheet_name=self.properties[self.__SHEET_NAME_PROPERTY],
                )[column_names]
            return pd.read_excel(
                self.properties[self.__REQUIRED_PATH_PROPERTY],
                sheet_name=self.properties[self.__SHEET_NAME_PROPERTY],
            )
        else:
            if usecols:
                return pd.read_excel(
                    self.properties[self.__REQUIRED_PATH_PROPERTY],
                    header=None,
                    usecols=usecols,
                    sheet_name=self.properties[self.__SHEET_NAME_PROPERTY],
                )
            return pd.read_excel(
                self.properties[self.__REQUIRED_PATH_PROPERTY],
                header=None,
                sheet_name=self.properties[self.__SHEET_NAME_PROPERTY],
            )

    def _write(self, data: Any):
        if isinstance(data, Dict) and all([isinstance(x, pd.DataFrame) for x in data.values()]):
            writer = pd.ExcelWriter(self.properties[self.__REQUIRED_PATH_PROPERTY])
            for key in data.keys():
                data[key].to_excel(writer, key, index=False)
            writer.save()
        else:
            pd.DataFrame(data).to_excel(self.properties[self.__REQUIRED_PATH_PROPERTY], index=False)

    def write_with_column_names(self, data: Any, columns: List[str] = None, job_id: Optional[JobId] = None):
        if not columns:
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame(data, columns=columns)
        df.to_excel(self.path, index=False)
        self.last_edition_date = datetime.now()
        if job_id:
            self.job_ids.append(job_id)
