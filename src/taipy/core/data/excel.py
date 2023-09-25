# Copyright 2023 Avaiga Private Limited
#
# Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
# the License. You may obtain a copy of the License at
#
#        http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on
# an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the
# specific language governing permissions and limitations under the License.

import os
from collections import defaultdict
from datetime import datetime, timedelta
from os.path import isfile
from typing import Any, Dict, List, Optional, Set, Tuple, Union

import modin.pandas as modin_pd
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from taipy.config.common.scope import Scope

from .._backup._backup import _replace_in_backup_file
from .._entity._reload import _self_reload
from .._version._version_manager_factory import _VersionManagerFactory
from ..exceptions.exceptions import ExposedTypeLengthMismatch, NonExistingExcelSheet, SheetNameLengthMismatch
from ..job.job_id import JobId
from ._abstract_file import _AbstractFileDataNode
from ._abstract_tabular import _AbstractTabularDataNode
from .data_node import DataNode
from .data_node_id import DataNodeId, Edit


class ExcelDataNode(DataNode, _AbstractFileDataNode, _AbstractTabularDataNode):
    """Data Node stored as an Excel file.

    The Excel file format is _xlsx_.

    Attributes:
        config_id (str): Identifier of this data node configuration. It must be a valid Python
            identifier.
        scope (Scope^): The scope of this data node.
        id (str): The unique identifier of this data node.
        name (str): A user-readable name of this data node.
        owner_id (str): The identifier of the owner (sequence_id, scenario_id, cycle_id) or
            `None`.
        parent_ids (Optional[Set[str]]): The identifiers of the parent tasks or `None`.
        last_edit_date (datetime): The date and time of the last modification.
        edits (List[Edit^]): The ordered list of edits for that job.
        version (str): The string indicates the application version of the data node to instantiate. If not provided,
            the current version is used.
        validity_period (Optional[timedelta]): The duration implemented as a timedelta since the last edit date for
            which the data node can be considered up-to-date. Once the validity period has passed, the data node is
            considered stale and relevant tasks will run even if they are skippable (see the
            [Task management page](../core/entities/task-mgt.md) for more details).
            If _validity_period_ is set to `None`, the data node is always up-to-date.
        edit_in_progress (bool): True if a task computing the data node has been submitted
            and not completed yet. False otherwise.
        editor_id (Optional[str]): The identifier of the user who is currently editing the data node.
        editor_expiration_date (Optional[datetime]): The expiration date of the editor lock.
        path (str): The path to the Excel file.
        properties (dict[str, Any]): A dictionary of additional properties. The _properties_
            must have a _"default_path"_ or _"path"_ entry with the path of the Excel file:

            - _"default_path"_ `(str)`: The path of the Excel file.\n
            - _"has_header"_ `(bool)`: If True, indicates that the Excel file has a header.\n
            - _"sheet_name"_ `(Union[List[str], str])`: The list of sheet names to be used. This
                can be a unique name.\n
            - _"exposed_type"_: The exposed type of the data read from Excel file. The default value is `pandas`.\n
    """

    __STORAGE_TYPE = "excel"
    __EXPOSED_TYPE_PROPERTY = "exposed_type"
    __EXPOSED_TYPE_NUMPY = "numpy"
    __EXPOSED_TYPE_PANDAS = "pandas"
    __EXPOSED_TYPE_MODIN = "modin"
    __VALID_STRING_EXPOSED_TYPES = [__EXPOSED_TYPE_PANDAS, __EXPOSED_TYPE_MODIN, __EXPOSED_TYPE_NUMPY]
    __PATH_KEY = "path"
    __DEFAULT_DATA_KEY = "default_data"
    __DEFAULT_PATH_KEY = "default_path"
    __HAS_HEADER_PROPERTY = "has_header"
    __SHEET_NAME_PROPERTY = "sheet_name"
    _REQUIRED_PROPERTIES: List[str] = []

    def __init__(
        self,
        config_id: str,
        scope: Scope,
        id: Optional[DataNodeId] = None,
        name: Optional[str] = None,
        owner_id: Optional[str] = None,
        parent_ids: Optional[Set[str]] = None,
        last_edit_date: Optional[datetime] = None,
        edits: List[Edit] = None,
        version: str = None,
        validity_period: Optional[timedelta] = None,
        edit_in_progress: bool = False,
        editor_id: Optional[str] = None,
        editor_expiration_date: Optional[datetime] = None,
        properties: Dict = None,
    ):
        if properties is None:
            properties = {}

        default_value = properties.pop(self.__DEFAULT_DATA_KEY, None)
        self._path = properties.get(self.__PATH_KEY, properties.get(self.__DEFAULT_PATH_KEY))
        properties[self.__PATH_KEY] = self._path

        if self.__SHEET_NAME_PROPERTY not in properties.keys():
            properties[self.__SHEET_NAME_PROPERTY] = None
        if self.__HAS_HEADER_PROPERTY not in properties.keys():
            properties[self.__HAS_HEADER_PROPERTY] = True
        if self.__EXPOSED_TYPE_PROPERTY not in properties.keys():
            properties[self.__EXPOSED_TYPE_PROPERTY] = self.__EXPOSED_TYPE_PANDAS
        self._check_exposed_type(properties[self.__EXPOSED_TYPE_PROPERTY], self.__VALID_STRING_EXPOSED_TYPES)

        super().__init__(
            config_id,
            scope,
            id,
            name,
            owner_id,
            parent_ids,
            last_edit_date,
            edits,
            version or _VersionManagerFactory._build_manager()._get_latest_version(),
            validity_period,
            edit_in_progress,
            editor_id,
            editor_expiration_date,
            **properties,
        )
        if not self._path:
            self._path = self._build_path(self.storage_type())
            properties[self.__PATH_KEY] = self._path

        if default_value is not None and not os.path.exists(self._path):
            self.write(default_value)

        if not self._last_edit_date and isfile(self._path):
            self._last_edit_date = datetime.now()

        self._TAIPY_PROPERTIES.update(
            {
                self.__EXPOSED_TYPE_PROPERTY,
                self.__PATH_KEY,
                self.__DEFAULT_PATH_KEY,
                self.__DEFAULT_DATA_KEY,
                self.__HAS_HEADER_PROPERTY,
                self.__SHEET_NAME_PROPERTY,
            }
        )

    @property  # type: ignore
    @_self_reload(DataNode._MANAGER_NAME)
    def path(self):
        return self._path

    @path.setter
    def path(self, value):
        tmp_old_path = self._path
        self._path = value
        self.properties[self.__PATH_KEY] = value
        _replace_in_backup_file(old_file_path=tmp_old_path, new_file_path=self._path)

    @classmethod
    def storage_type(cls) -> str:
        return cls.__STORAGE_TYPE

    @staticmethod
    def _check_exposed_type(exposed_type, valid_string_exposed_types):
        if isinstance(exposed_type, str):
            _AbstractTabularDataNode._check_exposed_type(exposed_type, valid_string_exposed_types)
        elif isinstance(exposed_type, list):
            for t in exposed_type:
                _AbstractTabularDataNode._check_exposed_type(t, valid_string_exposed_types)
        elif isinstance(exposed_type, dict):
            for t in exposed_type.values():
                _AbstractTabularDataNode._check_exposed_type(t, valid_string_exposed_types)

    def _read(self):
        if self.properties[self.__EXPOSED_TYPE_PROPERTY] == self.__EXPOSED_TYPE_PANDAS:
            return self._read_as_pandas_dataframe()
        if self.properties[self.__EXPOSED_TYPE_PROPERTY] == self.__EXPOSED_TYPE_MODIN:
            return self._read_as_modin_dataframe()
        if self.properties[self.__EXPOSED_TYPE_PROPERTY] == self.__EXPOSED_TYPE_NUMPY:
            return self._read_as_numpy()
        return self._read_as()

    def __sheet_name_to_list(self, properties):
        if properties[self.__SHEET_NAME_PROPERTY]:
            sheet_names = properties[self.__SHEET_NAME_PROPERTY]
        else:
            excel_file = load_workbook(properties[self.__PATH_KEY])
            sheet_names = excel_file.sheetnames
            excel_file.close()
        return sheet_names if isinstance(sheet_names, (List, Set, Tuple)) else [sheet_names]

    def _read_as(self):
        excel_file = load_workbook(self._path)
        exposed_type = self.properties[self.__EXPOSED_TYPE_PROPERTY]
        work_books = defaultdict()
        sheet_names = excel_file.sheetnames
        provided_sheet_names = self.__sheet_name_to_list(self.properties)

        for sheet_name in provided_sheet_names:
            if sheet_name not in sheet_names:
                raise NonExistingExcelSheet(sheet_name, self._path)

        if isinstance(exposed_type, List):
            if len(provided_sheet_names) != len(self.properties[self.__EXPOSED_TYPE_PROPERTY]):
                raise ExposedTypeLengthMismatch(
                    f"Expected {len(provided_sheet_names)} exposed types, got "
                    f"{len(self.properties[self.__EXPOSED_TYPE_PROPERTY])}"
                )

        for i, sheet_name in enumerate(provided_sheet_names):
            work_sheet = excel_file[sheet_name]
            sheet_exposed_type = exposed_type

            if not isinstance(sheet_exposed_type, str):
                if isinstance(exposed_type, dict):
                    sheet_exposed_type = exposed_type.get(sheet_name, self.__EXPOSED_TYPE_PANDAS)
                elif isinstance(exposed_type, List):
                    sheet_exposed_type = exposed_type[i]

                if isinstance(sheet_exposed_type, str):
                    if sheet_exposed_type == self.__EXPOSED_TYPE_NUMPY:
                        work_books[sheet_name] = self._read_as_pandas_dataframe(sheet_name).to_numpy()
                    elif sheet_exposed_type == self.__EXPOSED_TYPE_PANDAS:
                        work_books[sheet_name] = self._read_as_pandas_dataframe(sheet_name)
                    continue

            res = list()
            for row in work_sheet.rows:
                res.append([col.value for col in row])
            if self.properties[self.__HAS_HEADER_PROPERTY] and res:
                header = res.pop(0)
                for i, row in enumerate(res):
                    res[i] = sheet_exposed_type(**dict([[h, r] for h, r in zip(header, row)]))
            else:
                for i, row in enumerate(res):
                    res[i] = sheet_exposed_type(*row)
            work_books[sheet_name] = res

        excel_file.close()

        if len(provided_sheet_names) == 1:
            return work_books[provided_sheet_names[0]]
        return work_books

    def _read_as_numpy(self):
        sheets = self._read_as_pandas_dataframe()
        if isinstance(sheets, dict):
            return {sheet_name: df.to_numpy() for sheet_name, df in sheets.items()}
        return sheets.to_numpy()

    def _do_read_excel(self, engine, sheet_names, kwargs) -> pd.DataFrame:
        df = pd.read_excel(
            self._path,
            sheet_name=sheet_names,
            **kwargs,
        )
        # We are using pandas to load modin dataframes because of a modin issue
        # https://github.com/modin-project/modin/issues/4924
        if engine == "modin":
            if isinstance(df, dict):  # Check if it s a multiple sheet Excel file
                for key, value in df.items():
                    df[key] = modin_pd.DataFrame(value)
                return df
            return modin_pd.DataFrame(df)
        return df

    def __get_sheet_names_and_header(self, sheet_names):
        kwargs: Dict[str, Any] = {}
        if sheet_names is None:
            sheet_names = self.properties[self.__SHEET_NAME_PROPERTY]
        if not self.properties[self.__HAS_HEADER_PROPERTY]:
            kwargs["header"] = None
        return sheet_names, kwargs

    def _read_as_pandas_dataframe(self, sheet_names=None) -> Union[Dict[Union[int, str], pd.DataFrame], pd.DataFrame]:
        sheet_names, kwargs = self.__get_sheet_names_and_header(sheet_names)
        try:
            return self._do_read_excel("pandas", sheet_names, kwargs)
        except pd.errors.EmptyDataError:
            return pd.DataFrame()

    def _read_as_modin_dataframe(
        self, sheet_names=None
    ) -> Union[Dict[Union[int, str], modin_pd.DataFrame], modin_pd.DataFrame]:
        sheet_names, kwargs = self.__get_sheet_names_and_header(sheet_names)
        try:
            if kwargs.get("header", None):
                return modin_pd.read_excel(
                    self._path,
                    sheet_name=sheet_names,
                    **kwargs,
                )
            else:
                return self._do_read_excel("modin", sheet_names, kwargs)
        except pd.errors.EmptyDataError:
            return modin_pd.DataFrame()

    def __write_excel_with_single_sheet(self, write_excel_fct, *args, **kwargs):
        sheet_name = self.properties.get(self.__SHEET_NAME_PROPERTY)
        if sheet_name:
            if not isinstance(sheet_name, str):
                if len(sheet_name) > 1:
                    raise SheetNameLengthMismatch
                else:
                    sheet_name = sheet_name[0]
            write_excel_fct(*args, **kwargs, sheet_name=sheet_name)
        else:
            write_excel_fct(*args, **kwargs)

    def __write_excel_with_multiple_sheets(self, data: Any, columns: List[str] = None):
        with pd.ExcelWriter(self._path) as writer:
            # Each key stands for a sheet name
            for key in data.keys():
                if isinstance(data[key], np.ndarray):
                    df = pd.DataFrame(data[key])
                else:
                    df = data[key]

                if columns:
                    data[key].columns = columns

                df.to_excel(writer, key, index=False)

    def _write(self, data: Any):
        if isinstance(data, Dict) and all(
            [isinstance(x, (pd.DataFrame, modin_pd.DataFrame, np.ndarray)) for x in data.values()]
        ):
            self.__write_excel_with_multiple_sheets(data)
        elif isinstance(data, (pd.DataFrame, modin_pd.DataFrame)):
            self.__write_excel_with_single_sheet(data.to_excel, self._path, index=False)
        else:
            self.__write_excel_with_single_sheet(pd.DataFrame(data).to_excel, self._path, index=False)

    def write_with_column_names(self, data: Any, columns: List[str] = None, job_id: Optional[JobId] = None):
        """Write a set of columns.

        Parameters:
            data (Any): The data to write.
            columns (List[str]): The list of column names to write.
            job_id (JobId^): An optional identifier of the writer.
        """
        if isinstance(data, Dict) and all(
            [isinstance(x, (pd.DataFrame, modin_pd.DataFrame, np.ndarray)) for x in data.values()]
        ):
            self.__write_excel_with_multiple_sheets(data, columns=columns)
        else:
            df = pd.DataFrame(data)
            if columns:
                df.columns = columns
            self.__write_excel_with_single_sheet(df.to_excel, self.path, index=False)
        self.track_edit(timestamp=datetime.now(), job_id=job_id)
